# -*- coding: utf-8 -*-
"""Build Suvvy knowledge base xlsx from suvvy-qa/*.md. SKIP follow-ups."""

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

QA = Path(__file__).resolve().parent
OUT = QA / "knowledge_base.xlsx"
SKIP = {"11-follow-ups-reference.md"}


def parse_qa(path: Path):
    text = path.read_text(encoding="utf-8")
    m_title = re.search(r"\*\*Title \(exact\):\*\* `([^`]+)`", text)
    m_body = re.search(r"```\n(.*)\n```", text, re.S)
    if not m_title or not m_body:
        raise ValueError(f"Cannot parse {path.name} — need Title (exact) and fenced body")
    return m_title.group(1), m_body.group(1).strip()


def main() -> None:
    entries = []
    for path in sorted(QA.glob("[0-9][0-9]-*.md")):
        if path.name in SKIP:
            print("skip", path.name)
            continue
        title, body = parse_qa(path)
        entries.append((title, body))
        print(path.name, "->", title, len(body), "chars")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["Title", "Search Title", "Content", "Used"])
    for title, content in entries:
        ws.append([title, title, content, True])
        ws.cell(ws.max_row, 3).alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(OUT)
    print("wrote", OUT, "rows", len(entries))


if __name__ == "__main__":
    main()
